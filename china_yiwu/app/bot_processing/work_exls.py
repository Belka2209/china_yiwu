from datetime import date
import os
import shutil
from pg.pg_query import get_data_in_table, get_data_in_table_all
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import telegram
TOKEN = os.getenv("TOKEN")

bot = telegram.Bot(TOKEN)

async def add_form_excl_doc(chat_id, id_order):
    '''Формирование файла Excel с данными заказа'''
    # Получаем данные заказа из БД
    order_data = await get_data_in_table_all("orders", "id_order", id_order)
    print("order_data", order_data)
    source_file = "/data/model.xlsx"
    if order_data[0].get("delivery_direction") == "Москва":
        pr1 = "M"
    elif order_data[0].get("delivery_direction") == "Уссурийск":
        pr1 = "U"
    elif order_data[0].get("delivery_direction") == "Казахстан":
        pr1 = "KZ"
    user_data = await get_data_in_table("users", "chat_id", chat_id)
    pr2 = user_data.get("phone")[-7:].replace("-", "")
    today = date.today()
    new_file_name = f"{pr1}_{pr2}_{id_order}_{today}.xlsx"
    target_dir = f"/data/orders/{chat_id}/"
    new_file = os.path.join(target_dir, new_file_name)
    # print("new_file", new_file)
    os.makedirs(target_dir, exist_ok=True)
    shutil.copy(source_file, new_file)
    wb = load_workbook(new_file)
    ws = wb.active
    ws["D2"] = user_data.get("fio")
    ws["D5"] = user_data.get("phone")
    ws["I6"] = user_data.get("tg_name")

    start_row = 8  # Начинаем с 8 строки (если шапка до 7 строки)
    for i, order in enumerate(order_data):
        current_row = start_row + i  # Увеличиваем строку для каждого заказа
        ws[f"C{current_row}"] = order.get("link_product")
        ws[f"H{current_row}"] = order.get("count")
        ws[f"I{current_row}"] = order.get("price")
        ws[f"J{current_row}"] = order.get("amount")
        for img_type, col in [
            ("img_product", "F"),
            ("img_color", "E"),
            ("img_size", "D"),
        ]:
            img_path = order.get(img_type)
            if img_path and os.path.exists(img_path):
                img = Image(img_path)
                img.width = 200
                img.height = 200
                ws.add_image(img, f"{col}{current_row}")

        amount = int(order.get("amount"))
        # print("amount", amount)
        if amount <= 1000:
            ws[f"K{current_row}"] = 7
        elif 1000 < amount <= 3000:
            ws[f"K{current_row}"] = 5
        else:
            ws[f"K{current_row}"] = 3

    wb.save(new_file)
    with open(new_file, 'rb') as file:
        await bot.send_document(
            # chat_id=chat_id,
            chat_id=464606401,
            document=file,
            caption=f"Сформирован новый заказ №{id_order}",
            filename=f"{new_file_name}"
        )
    return new_file_name